"""Ingestion pipeline.

Restores / fixes:
- ingest_test_cases(): parses the team's existing .xlsx test cases and stores them
  as Test_Case chunks linked to their issue key (the few-shot foundation - Req 2).
- content-hash change detection instead of mtime (Req 9.2): unchanged files skip.
- model-change detection (Req 10.5): if the store was built with a different
  embedding model, signal that a full re-index is required.
- lightweight default-on PII guard (Req 12.2)."""
from __future__ import annotations

import hashlib
import re
import time
from datetime import datetime, timezone
from pathlib import Path

from chunker import Chunker
from config import AppConfig
from embedder import EmbeddingService
from models import Chunk, IngestionResult, authority_for
from store import ChromaStore

SUPPORTED_TEXT = {".md", ".txt", ".json"}
_SSN = re.compile(r"\b\d{3}-\d{2}-\d{4}\b")
_BANK = re.compile(r"\b\d{9,18}\b")          # coarse: routing/account-length digit runs
_TC_NAME = re.compile(r"^([A-Z][A-Z0-9]+-\d+)_TC-\d+", re.IGNORECASE)


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def redact_pii(text: str) -> tuple[str, int]:
    n = 0
    def sub(p, repl, s):
        nonlocal n
        s2, c = p.subn(repl, s)
        n += c
        return s2
    text = sub(_SSN, "[REDACTED-SSN]", text)
    text = sub(_BANK, "[REDACTED-ACCT]", text)
    return text, n


class IngestionPipeline:
    def __init__(self, config: AppConfig, chunker: Chunker,
                 embedder: EmbeddingService, store: ChromaStore,
                 conflict_detector=None):
        self.config, self.chunker, self.embedder, self.store = config, chunker, embedder, store
        # Optional Phase C hook; when set and config.conflict_detection is on, each
        # ingest scans freshly stored chunks for similarity conflicts.
        self.conflict_detector = conflict_detector

    # --- model-change guard --------------------------------------------------
    def model_needs_reindex(self) -> bool:
        versions = {v for v in self.store.stored_model_versions() if v}
        return bool(versions) and versions != {self.embedder.model_version}

    # --- versioned store (Phase A state machine) -----------------------------
    def _store_versioned(self, source: str, document_id: str, content_hash: str,
                         text: str, category: str, extra: dict, module: str,
                         title: str, t0: float, redactions: int = 0,
                         force: bool = False) -> IngestionResult:
        """Option B: never delete on update. New doc -> v1. Unchanged -> skip.
        Changed -> flip prior latest to is_latest=False and insert the next version."""
        latest = self.store.latest_version_for_document(document_id)
        if not force and latest is not None and latest[1] == content_hash:
            return IngestionResult(True, source, document_id=document_id, version=latest[0],
                                   skipped_unchanged=True, redactions=redactions,
                                   duration_ms=int((time.time() - t0) * 1000))

        version = 1 if latest is None else latest[0] + 1
        if latest is not None:
            self.store.mark_not_latest(document_id)

        base = self._base_meta(source, title, content_hash, category, extra,
                               document_id, version, module)
        chunks = self.chunker.chunk_document(text, base)
        embeddings = self._embed_and_store(chunks)

        if self.conflict_detector is not None and self.config.conflict_detection and chunks:
            # Never let detection failures break ingestion.
            try:
                self.conflict_detector.scan(chunks, embeddings)
            except Exception:  # noqa: BLE001
                pass

        if self.config.max_versions_retained and self.config.max_versions_retained > 0:
            self.store.prune_old_versions(document_id, self.config.max_versions_retained)

        return IngestionResult(True, source, document_id=document_id, version=version,
                               chunks_created=len(chunks), redactions=redactions,
                               duration_ms=int((time.time() - t0) * 1000))

    # --- local files ---------------------------------------------------------
    def ingest_file(self, file_path: str, category: str = "auto",
                    module: str = "auto", force: bool = False) -> IngestionResult:
        t0 = time.time()
        p = Path(file_path)
        try:
            raw = p.read_bytes()
        except OSError as e:
            return IngestionResult(False, file_path, errors=[f"read failed: {e}"])

        content_hash = hashlib.sha256(raw).hexdigest()
        text = raw.decode("utf-8", errors="replace")
        redactions = 0
        if self.config.pii_guard:
            text, redactions = redact_pii(text)

        return self._store_versioned(
            str(p), self._document_id_for_file(p), content_hash, text,
            self._auto_category(p, category), {}, self._module_for_file(p, module),
            p.stem, t0, redactions, force,
        )

    def ingest_directory(self, dir_path: str, category: str = "auto",
                         module: str = "auto") -> list[IngestionResult]:
        results = []
        for f in Path(dir_path).rglob("*"):
            if f.suffix.lower() in SUPPORTED_TEXT:
                results.append(self.ingest_file(str(f), category, module))
            elif f.suffix.lower() == ".xlsx":
                results.extend(self.ingest_test_cases(str(f)))
        return results

    # --- historical test cases (was entirely missing) -----------------------
    def ingest_test_cases(self, xlsx_path: str) -> list[IngestionResult]:
        """Parse an existing test-case workbook into per-story Test_Case chunks."""
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        results = []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            issue_key = self._issue_key_for_sheet(ws.title, rows)
            text = self._rows_to_text(rows)
            content_hash = hashlib.sha256(
                f"{xlsx_path}:{ws.title}:{text}".encode()).hexdigest()
            source = f"{xlsx_path}#{ws.title}"
            extra = {"issue_key": issue_key} if issue_key else {}
            results.append(self._store_versioned(
                source, source, content_hash, text, "Test_Case", extra,
                "default", ws.title, time.time(),
            ))
        return results

    # --- jira ----------------------------------------------------------------
    def ingest_jira_story(self, issue_key: str, jira_client) -> IngestionResult:
        t0 = time.time()
        try:
            story = jira_client.get_issue(issue_key)
        except Exception as e:
            return IngestionResult(False, f"jira:{issue_key}", errors=[str(e)])
        return self._ingest_story_dict(story, story.get("epic_key", ""), t0)

    def ingest_jira_epic(self, epic_key: str, jira_client, fail_fast: bool = False) -> IngestionResult:
        """Best-effort by default: collect per-story failures and report (Req 3.5)."""
        t0 = time.time()
        try:
            stories = jira_client.get_epic_stories(epic_key)
        except Exception as e:
            return IngestionResult(False, f"jira:{epic_key}", errors=[str(e)])

        total_created = total_deleted = total_redactions = 0
        errors: list[str] = []
        for story in stories:
            issue_key = story.get("issue_key", "")
            if not issue_key:
                continue
            try:
                r = self._ingest_story_dict(story, epic_key, time.time())
                total_created += r.chunks_created
                total_deleted += r.chunks_deleted
                total_redactions += r.redactions
                if r.errors:
                    errors.extend(r.errors)
            except Exception as e:
                if fail_fast:
                    return IngestionResult(False, f"jira:{epic_key}",
                                           errors=[f"{issue_key}: {e}"])
                errors.append(f"{issue_key}: {e}")

        return IngestionResult(
            success=not errors or total_created > 0,
            source_path=f"jira:{epic_key}",
            chunks_created=total_created, chunks_deleted=total_deleted,
            redactions=total_redactions, errors=errors,
            duration_ms=int((time.time() - t0) * 1000),
        )

    def sync(self, dir_path: str) -> list[IngestionResult]:
        """Ingest new/changed files and delete chunks for removed sources."""
        results: list[IngestionResult] = []
        fs_paths: set[str] = set()

        for f in Path(dir_path).rglob("*"):
            suffix = f.suffix.lower()
            if suffix in SUPPORTED_TEXT:
                fs_paths.add(str(f))
                results.append(self.ingest_file(str(f)))
            elif suffix == ".xlsx":
                fs_paths.add(str(f))
                results.extend(self.ingest_test_cases(str(f)))

        # Purge chunks whose backing file no longer exists (skip jira: sources)
        for source in self.store.get_all_sources():
            if source.startswith("jira:"):
                continue
            base = source.split("#")[0]   # handles "path.xlsx#SheetName" form
            if base not in fs_paths:
                deleted = self.store.delete_by_source(source)
                if deleted:
                    results.append(IngestionResult(True, source, chunks_deleted=deleted))

        return results

    # --- shared jira helper --------------------------------------------------
    def _ingest_story_dict(self, story: dict, epic_key: str, t0: float) -> IngestionResult:
        """Build text from a normalized story dict, hash-check, chunk, and store."""
        issue_key = story.get("issue_key", "")
        parts = [
            f"# {story.get('summary', '')}",
            story.get("description", ""),
            "## Acceptance Criteria",
            story.get("acceptance_criteria", ""),
            "## Definition of Done",
            story.get("definition_of_done", ""),
        ]
        comments = story.get("comments") or []
        if comments:
            parts.append("## Comments")
            parts.extend(c for c in comments if c)
        text = "\n\n".join(p for p in parts if p and p.strip())

        redactions = 0
        if self.config.pii_guard:
            text, redactions = redact_pii(text)

        content_hash = hashlib.sha256(text.encode()).hexdigest()
        source = f"jira:{issue_key}"
        meta = {
            "issue_key": issue_key, "epic_key": epic_key,
            "assignee": story.get("assignee", ""), "status": story.get("status", ""),
        }
        return self._store_versioned(
            source, source, content_hash, text, "Story", meta, "default",
            story.get("summary") or issue_key, t0, redactions,
        )

    # --- resolutions (Phase C) ----------------------------------------------
    def ingest_resolution(self, text: str, resolution_id: str, module: str = "default",
                          extra: dict | None = None) -> IngestionResult:
        """Re-ingest a human-approved resolution as a high-authority chunk so
        retrieval prioritizes it over the original conflicting sources."""
        content_hash = hashlib.sha256(text.encode()).hexdigest()
        source = f"resolution:{resolution_id}"
        return self._store_versioned(
            source, source, content_hash, text, "Business_Resolution",
            extra or {}, module, f"Resolution {resolution_id}", time.time(),
        )

    # --- helpers -------------------------------------------------------------
    def _embed_and_store(self, chunks: list[Chunk]) -> list[list[float]]:
        if not chunks:
            return []
        embeddings = self.embedder.embed_texts([c.text for c in chunks])
        self.store.upsert_chunks(chunks, embeddings)
        return embeddings

    def _base_meta(self, source, title, content_hash, category, extra,
                   document_id, version, module) -> dict:
        return {
            "source_path": source, "document_title": title, "content_hash": content_hash,
            "category": category, "ingestion_timestamp": _now(),
            "embedding_model": self.embedder.model_name,
            "model_version": self.embedder.model_version,
            "document_id": document_id, "version": version, "is_latest": True,
            "module": module, "authority_score": authority_for(category),
            "metadata": extra,
        }

    def _document_id_for_file(self, p: Path) -> str:
        """Stable logical identity across versions: POSIX path relative to
        knowledge_dir, else the absolute POSIX path if outside it."""
        try:
            return p.resolve().relative_to(Path(self.config.knowledge_dir).resolve()).as_posix()
        except (ValueError, OSError):
            return p.resolve().as_posix()

    def _module_for_file(self, p: Path, module: str) -> str:
        """Explicit param wins; else the first folder under knowledge_dir; else default."""
        if module and module != "auto":
            return module
        try:
            rel = p.resolve().relative_to(Path(self.config.knowledge_dir).resolve())
            if len(rel.parts) > 1:
                return rel.parts[0]
        except (ValueError, OSError):
            pass
        return "default"

    @staticmethod
    def _auto_category(path: Path, category: str) -> str:
        if category != "auto":
            return category
        name = path.name.lower()
        if "mom" in name or "minutes" in name:
            return "MOM"
        if "error" in name or "code" in name:
            return "Error_Code"
        if "flow" in name or "ui" in name:
            return "UI_Flow"
        return "Business_Rule"

    @staticmethod
    def _issue_key_for_sheet(sheet_title: str, rows) -> str | None:
        m = re.match(r"([A-Z][A-Z0-9]+-\d+)", sheet_title)
        if m:
            return m.group(1)
        for row in rows[1:]:
            for cell in row:
                if isinstance(cell, str):
                    mm = _TC_NAME.match(cell)
                    if mm:
                        return mm.group(1)
        return None

    @staticmethod
    def _rows_to_text(rows) -> str:
        return "\n".join(
            " | ".join("" if c is None else str(c) for c in row) for row in rows
        )
