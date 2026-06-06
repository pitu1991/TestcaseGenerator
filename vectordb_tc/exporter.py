"""Excel exporter. Writes the team-standard template. Called only AFTER the
validator passes (or with already-validated cases)."""
from __future__ import annotations

from pathlib import Path

from models import TestCase


class ExcelExporter:
    HEADERS = ["Name", "Description", "Precondition", "Test Step #",
               "Test Step Description", "Expected", "Assigned To",
               "Requirement Id", "Status", "Type", "Workplace Capability",
               "Priority", "Application"]
    COLUMN_WIDTHS = {"A": 70, "B": 40, "C": 40, "D": 8, "E": 55, "F": 50, "G": 12, "H": 12}
    DEFAULTS = {"Status": "Not Run", "Type": "Manual", "Application": "CORE"}

    def export(self, cases: list[TestCase], issue_key: str, output_dir: str) -> str:
        return self.export_multi({issue_key: cases}, output_dir)

    def export_multi(self, stories: dict[str, list[TestCase]], output_dir: str) -> str:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)
        for issue_key, cases in stories.items():
            ws = wb.create_sheet(title=issue_key[:31])
            ws.append(self.HEADERS)
            for tc in cases:
                first = True
                steps = tc.steps or [None]
                for i, step in enumerate(steps, 1):
                    ws.append([
                        tc.name if first else "",
                        tc.description if first else "",
                        tc.precondition if first else "",
                        str(step.step_number) if step else str(i),
                        step.description if step else "",
                        step.expected_result if step else "",
                        tc.assigned_to if first else "",
                        tc.requirement_id if first else "",
                        tc.status if first else "",
                        tc.type if first else "",
                        tc.workplace_capability if first else "",
                        tc.priority if first else "",
                        tc.application if first else "",
                    ])
                    first = False
            for col, width in self.COLUMN_WIDTHS.items():
                ws.column_dimensions[col].width = width
            _ = get_column_letter  # available for any extra dynamic columns

        Path(output_dir).mkdir(parents=True, exist_ok=True)
        keys = "_".join(stories.keys())[:60]
        out = str(Path(output_dir) / f"{keys}_testcases.xlsx")
        wb.save(out)
        return out
